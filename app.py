"""My Lending — web application entry point.

Run with:
    python app.py
Then open http://127.0.0.1:5000 in a browser. The first visitor is asked
to create the admin account; every screen after that requires sign-in.
"""

import csv
import io
import secrets
from datetime import date, datetime
from functools import wraps
from pathlib import Path

from flask import (
    Flask, render_template, request, redirect, url_for,
    session, flash, send_file, abort, g
)

from database import LendingDatabase

BASE_DIR = Path(__file__).resolve().parent
INSTANCE_DIR = BASE_DIR / "instance"
INSTANCE_DIR.mkdir(exist_ok=True)
SECRET_KEY_FILE = INSTANCE_DIR / "secret_key"
DB_PATH = BASE_DIR / "my_lending.db"

if not SECRET_KEY_FILE.exists():
    SECRET_KEY_FILE.write_text(secrets.token_hex(32))

app = Flask(__name__)
app.config["SECRET_KEY"] = SECRET_KEY_FILE.read_text().strip()

db = LendingDatabase(str(DB_PATH))


def peso(value):
    return f"₱{float(value):,.2f}"


app.jinja_env.filters["peso"] = peso


# --------------------------------------------------------------------- auth
def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not db.has_admin():
            return redirect(url_for("setup"))
        if not session.get("user_id"):
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


@app.before_request
def load_current_user():
    g.user = None
    user_id = session.get("user_id")
    if user_id:
        g.user = db.get_user(user_id)


@app.context_processor
def inject_user():
    return {"current_user": g.get("user")}


@app.route("/setup", methods=["GET", "POST"])
def setup():
    if db.has_admin():
        return redirect(url_for("login"))
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        confirm = request.form.get("confirm", "")
        if not username or len(username) < 3:
            flash("Choose a username with at least 3 characters.", "error")
        elif len(password) < 8:
            flash("Choose a password with at least 8 characters.", "error")
        elif password != confirm:
            flash("Passwords do not match.", "error")
        else:
            db.create_user(username, password)
            flash("Admin account created. Please sign in.", "success")
            return redirect(url_for("login"))
    return render_template("setup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if not db.has_admin():
        return redirect(url_for("setup"))
    if session.get("user_id"):
        return redirect(url_for("dashboard"))
    if request.method == "POST":
        username = request.form.get("username", "")
        password = request.form.get("password", "")
        user = db.verify_user(username, password)
        if user:
            session.clear()
            session["user_id"] = user["id"]
            session.permanent = True
            return redirect(request.args.get("next") or url_for("dashboard"))
        flash("Incorrect username or password.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/account", methods=["GET", "POST"])
@login_required
def account():
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        if not db.verify_user(g.user["username"], current):
            flash("Current password is incorrect.", "error")
        elif len(new) < 8:
            flash("Choose a new password with at least 8 characters.", "error")
        elif new != confirm:
            flash("New passwords do not match.", "error")
        else:
            db.update_password(g.user["id"], new)
            flash("Password updated.", "success")
    return render_template("account.html")


# ---------------------------------------------------------------- dashboard
@app.route("/")
@login_required
def dashboard():
    summary = db.summary()
    recent = db.loans()[:7]
    return render_template("dashboard.html", summary=summary, recent=recent)


# ---------------------------------------------------------------- new loan
@app.route("/loans/new", methods=["GET", "POST"])
@login_required
def new_loan():
    if request.method == "POST":
        try:
            borrower = request.form.get("borrower", "").strip()
            contact = request.form.get("contact", "").strip()
            principal = float(request.form.get("principal", ""))
            rate = float(request.form.get("interest", ""))
            months = int(request.form.get("months", ""))
            if not borrower or principal <= 0 or rate < 0 or months <= 0:
                raise ValueError
            db.add_loan(borrower, contact, principal, rate, months)
            flash(f"Loan saved. Total interest: {peso(principal * rate / 100 * months)}", "success")
            return redirect(url_for("dashboard"))
        except ValueError:
            flash("Enter a borrower, positive principal and months, and a valid interest rate.", "error")
    return render_template("new_loan.html")


# ---------------------------------------------------------------- borrowers
@app.route("/borrowers")
@login_required
def borrowers():
    term = request.args.get("q", "")
    rows = db.search_loans(term)
    return render_template("borrowers.html", rows=rows, term=term)


@app.route("/borrowers/<int:loan_id>")
@login_required
def borrower_detail(loan_id):
    loan = db.loan_detail(loan_id)
    if loan is None:
        abort(404)
    history = db.loan_payments(loan_id)
    return render_template("borrower_detail.html", loan=loan, history=history)


@app.route("/borrowers/<int:loan_id>/rename", methods=["POST"])
@login_required
def rename_borrower(loan_id):
    loan = db.loan_detail(loan_id)
    if loan is None:
        abort(404)
    name = request.form.get("borrower", "").strip()
    if not name:
        flash("Enter the borrower's name.", "error")
    else:
        db.update_borrower_name(loan_id, name)
        flash("Borrower name updated.", "success")
    return redirect(request.form.get("redirect_to") or url_for("borrower_detail", loan_id=loan_id))


@app.route("/borrowers/<int:loan_id>/delete", methods=["POST"])
@login_required
def delete_borrower(loan_id):
    loan = db.loan_detail(loan_id)
    if loan is None:
        abort(404)
    db.delete_loan(loan_id)
    flash(f"{loan['borrower']}'s loan and payment history were deleted.", "success")
    return redirect(url_for("borrowers"))


# ----------------------------------------------------------------- payments
@app.route("/payments")
@login_required
def payments():
    return render_template("payments.html", rows=db.payments())


@app.route("/payments/new", methods=["GET", "POST"])
@login_required
def new_payment():
    choices = db.loan_choices()
    if request.method == "POST":
        try:
            loan_id = int(request.form.get("loan_id"))
            amount = float(request.form.get("amount"))
            paid_at = request.form.get("paid_at") or date.today().isoformat()
            note = request.form.get("note", "").strip()
            if amount <= 0:
                raise ValueError
            db.add_payment(loan_id, amount, paid_at, note)
            flash("Payment saved.", "success")
            return redirect(url_for("payments"))
        except (ValueError, TypeError):
            flash("Enter a valid loan, amount, and date.", "error")
    if not choices:
        flash("Create an active loan before recording a payment.", "error")
        return redirect(url_for("payments"))
    return render_template("new_payment.html", choices=choices, today=date.today().isoformat())


# ------------------------------------------------------------------ export
@app.route("/borrowers/export/<fmt>")
@login_required
def export_data(fmt):
    rows = db.search_loans(request.args.get("q", ""))
    headers = ["Loan ID", "Borrower", "Contact", "Principal (PHP)", "Monthly Interest (%)",
               "Months", "Total Interest (PHP)", "Total Payment (PHP)", "Payments Received (PHP)",
               "Running Balance (PHP)", "Status", "Created At"]
    values = [[r["id"], r["borrower"], r["contact"], r["principal"], r["monthly_interest"],
               r["months"], r["total_interest"], r["total_payable"], r["paid"], r["balance"],
               r["status"], r["created_at"]] for r in rows]

    if fmt == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)
        writer.writerows(values)
        mem = io.BytesIO(buffer.getvalue().encode("utf-8-sig"))
        return send_file(mem, mimetype="text/csv", as_attachment=True,
                          download_name="my_lending_borrowers.csv")

    if fmt == "excel":
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            flash("Excel export needs a package. Install it once with: pip install openpyxl", "error")
            return redirect(url_for("borrowers"))
        book = Workbook()
        sheet = book.active
        sheet.title = "Borrowers"
        sheet.append(headers)
        for row in values:
            sheet.append(row)
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="102A43")
        for column in sheet.columns:
            sheet.column_dimensions[column[0].column_letter].width = min(
                max(len(str(c.value or "")) for c in column) + 2, 28)
        for row in sheet.iter_rows(min_row=2, min_col=4, max_col=10):
            for cell in row:
                cell.number_format = "#,##0.00"
        mem = io.BytesIO()
        book.save(mem)
        mem.seek(0)
        return send_file(mem, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                          as_attachment=True, download_name="my_lending_borrowers.xlsx")

    abort(404)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5000, debug=False)
